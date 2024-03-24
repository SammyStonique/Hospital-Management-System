import os
from django.shortcuts import render,get_object_or_404
from django.http import HttpResponse, JsonResponse
from .models import *
from users.models import Manager
from company.models import Company
from .serializers import *
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework import generics,status
from rest_framework.response import Response
#Pagination
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
#Pdf 
import jinja2
import pdfkit
#Excel
import xlwt
#CSV
import csv

import json
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt

import uuid
from datetime import datetime, timedelta

#Data Import
from openpyxl import load_workbook

# Create your views here.

        #PAGINATION

class BasePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 1000

class DefaultPagination(PageNumberPagination):
    page_size = 1000
    page_size_query_param = 'page_size'
    max_page_size = 1000

            #CLIENT CATEGORIES VIEWS
    
class ClientCategoryViewSet(viewsets.ModelViewSet):
    queryset = ClientCategory.objects.all()
    serializer_class = ClientCategorySerializer
    pagination_class = BasePagination

class ClientCategoryList(generics.ListCreateAPIView):
    queryset = ClientCategory.objects.all()
    serializer_class = ClientCategorySerializer
    pagination_class = DefaultPagination

class ClientCategoryDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = ClientCategory.objects.all()
    serializer_class = ClientCategorySerializer
        

@csrf_exempt
@api_view(['POST'])
def createClientCategory(request):
    company_id = request.data.get("company")
    company_uuid = uuid.UUID(company_id)
    company = get_object_or_404(Company, company_id=company_uuid)
    serializer = ClientCategorySerializer(data=request.data)

    if serializer.is_valid():
        serializer.save(company=company)

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def getClientCategories(request):
    category_id = request.data.get("category")
    company_id = request.data.get("company")

    if category_id is not None:
        company_uuid = uuid.UUID(company_id)
        category_uuid = uuid.UUID(category_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        category = ClientCategory.objects.get(company=company, category_id=category_uuid)

        serializer = ClientCategorySerializer(category)
        return Response(serializer.data)

    else:
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        categories = ClientCategory.objects.filter(company=company)

        serializer = ClientCategorySerializer(categories, many=True)
        return Response(serializer.data)


@csrf_exempt
@api_view(['PUT'])
def updateClientCategory(request):
    category_id = request.data.get("category")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    category_uuid = uuid.UUID(category_id)
    category = ClientCategory.objects.get(company=company_id,category_id=category_uuid)
    
    serializer = ClientCategorySerializer(category, data=request.data)

    if serializer.is_valid():
        serializer.save()

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def deleteClientCategory(request):
    category_id = request.data.get("category")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    category_uuid = uuid.UUID(category_id)
    category = ClientCategory.objects.get(company=company_id,category_id=category_uuid)

    category.delete() 
    message = {'msg':"The category has been succesfully deleted"} 
    return Response(message)


@csrf_exempt   
def generate_client_categories_pdf(request):
    categoryList = []
    data = json.loads(request.body)
    category_name = data['category_name']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    client_categories = ClientCategory.objects.filter(company=company_uuid)

    categories = client_categories.filter(Q(name__icontains=category_name) )

    for cat in categories:
        obj = {
            "category_id": cat.category_id,
            "category_name": cat.name,

        }
        categoryList.append(obj)

    context = {"categories":categoryList}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/templates/financial_accounts_chart_of_accounts')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('clientCategoryPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Client Category.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/static/financial_accounts_chart_of_accounts/journalsPDF.css")

    path = 'Client Category.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Client Category.pdf'
    pdf.close()
    os.remove("Client Category.pdf")  # remove the locally created pdf file.
    return response

@csrf_exempt
def generate_client_categories_excel(request):
    categoryList = []
    data = json.loads(request.body)
    category_name = data['category_name']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    client_categories = ClientCategory.objects.filter(company=company_uuid)

    categories = client_categories.filter(Q(name__icontains=category_name) )

    for cat in categories:
        obj = {
            "category_id": cat.category_id,
            "category_name": cat.name,

        }
        categoryList.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Client Categories.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Client Categories")

    row_num = 0
    columns = ['Client Category Name']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style=style1)

    for cat in categoryList:
        row_num += 1
        row = [cat['category_name']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@csrf_exempt
def generate_client_categories_csv(request):
    categoryList = []
    data = json.loads(request.body)
    category_name = data['category_name']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    client_categories = ClientCategory.objects.filter(company=company_uuid)

    categories = client_categories.filter(Q(name__icontains=category_name) )

    for cat in categories:
        obj = {
            "category_id": cat.category_id,
            "category_name": cat.name,

        }
        categoryList.append(obj)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Client Category.csv'

    writer = csv.writer(response)
    writer.writerow([ 'Client Category Name'])

    for cat in categoryList:
        writer.writerow([cat['category_name']])
    return response


            #LEDGERS VIEWS
    
class LedgerViewSet(viewsets.ModelViewSet):
    queryset = Ledger.objects.all()
    serializer_class = LedgerSerializer
    pagination_class = BasePagination

class LedgerList(generics.ListCreateAPIView):
    queryset = Ledger.objects.all()
    serializer_class = LedgerSerializer
    pagination_class = DefaultPagination

class LedgerDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = Ledger.objects.all()
    serializer_class = LedgerSerializer
        

@csrf_exempt
@api_view(['POST'])
def createLedger(request):
    company_id = request.data.get("company")
    company_uuid = uuid.UUID(company_id)
    company = get_object_or_404(Company, company_id=company_uuid)
    serializer = LedgerSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save(company=company)

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def getLedgers(request):
    ledger_id = request.data.get("ledger")
    company_id = request.data.get("company")
    ledger_type = request.data.get("ledger_type")

    if ledger_id is not None:
        company_uuid = uuid.UUID(company_id)
        ledger_uuid = uuid.UUID(ledger_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        ledger = Ledger.objects.get(company=company, ledger_id=ledger_uuid)

        serializer = LedgerSerializer(ledger)
        return Response(serializer.data)
    
    elif ledger_type is not None:
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        ledgers = Ledger.objects.filter(company=company, ledger_type=ledger_type)

        serializer = LedgerSerializer(ledgers, many=True)
        return Response(serializer.data)

    else:
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        ledgers = Ledger.objects.filter(company=company)

        serializer = LedgerSerializer(ledgers, many=True)
        return Response(serializer.data)


@csrf_exempt
@api_view(['PUT'])
def updateLedger(request):
    ledger_id = request.data.get("ledger")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    ledger_uuid = uuid.UUID(ledger_id)
    ledger = Ledger.objects.get(company=company_id,ledger_id=ledger_uuid)
    
    serializer = LedgerSerializer(ledger, data=request.data)

    if serializer.is_valid():
        serializer.save()

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def deleteLedger(request):
    ledger_id = request.data.get("ledger")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    ledger_uuid = uuid.UUID(ledger_id)
    ledger = Ledger.objects.get(company=company_id,ledger_id=ledger_uuid)

    ledger.delete() 
    message = {'msg':"The ledger has been succesfully deleted"} 
    return Response(message)


@csrf_exempt   
def generate_ledgers_pdf(request):
    chartOfAccountsList = []
    data = json.loads(request.body)
    ledger_code = data['ledger_code']
    ledger_name = data['ledger_name']
    financial_statement = data['financial_statement']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_ledgers = Ledger.objects.filter(company=company_uuid)

    ledgers = company_ledgers.filter(Q(ledger_name__icontains=ledger_name) & Q(ledger_code__icontains=ledger_code))

    if financial_statement:
        ledgers = ledgers.filter(financial_statement = financial_statement)

    for led in ledgers:
        obj = {
            "ledger_id": led.ledger_id,
            "ledger_code": led.ledger_code,
            "ledger_name": led.ledger_name,
            "ledger_type": led.ledger_type,
            "financial_statement": led.financial_statement,
            "balance": led.balance,

        }
        chartOfAccountsList.append(obj)

    context = {"ledgers":chartOfAccountsList}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/templates/financial_accounts_chart_of_accounts')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('ledgerPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Chart Of Accounts.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/static/financial_accounts_chart_of_accounts/clientCategoryPDF.css")

    path = 'Chart Of Accounts.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Chart Of Accounts.pdf'
    pdf.close()
    os.remove("Chart Of Accounts.pdf")  # remove the locally created pdf file.
    return response

@csrf_exempt
def generate_ledgers_excel(request):
    chartOfAccountsList = []
    data = json.loads(request.body)
    ledger_code = data['ledger_code']
    ledger_name = data['ledger_name']
    financial_statement = data['financial_statement']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_ledgers = Ledger.objects.filter(company=company_uuid)

    ledgers = company_ledgers.filter(Q(ledger_name__icontains=ledger_name) & Q(ledger_code__icontains=ledger_code))

    if financial_statement:
        ledgers = ledgers.filter(financial_statement = financial_statement)

    for led in ledgers:
        obj = {
            "ledger_id": led.ledger_id,
            "ledger_code": led.ledger_code,
            "ledger_name": led.ledger_name,
            "ledger_type": led.ledger_type,
            "financial_statement": led.financial_statement,
            "balance": led.balance,

        }
        chartOfAccountsList.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Chart Of Accounts.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Chart Of Accounts")

    row_num = 0
    columns = ['Client Category Name']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style=style1)

    for led in chartOfAccountsList:
        row_num += 1
        row = [led['category_name']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@csrf_exempt
def generate_ledgers_csv(request):
    chartOfAccountsList = []
    data = json.loads(request.body)
    ledger_code = data['ledger_code']
    ledger_name = data['ledger_name']
    financial_statement = data['financial_statement']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_ledgers = Ledger.objects.filter(company=company_uuid)

    ledgers = company_ledgers.filter(Q(ledger_name__icontains=ledger_name) & Q(ledger_code__icontains=ledger_code))

    if financial_statement:
        ledgers = ledgers.filter(financial_statement = financial_statement)

    for led in ledgers:
        obj = {
            "ledger_id": led.ledger_id,
            "ledger_code": led.ledger_code,
            "ledger_name": led.ledger_name,
            "ledger_type": led.ledger_type,
            "financial_statement": led.financial_statement,
            "balance": led.balance,

        }
        chartOfAccountsList.append(obj)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Chart Of Accounts.csv'

    writer = csv.writer(response)
    writer.writerow([ 'Client Category Name'])

    for led in chartOfAccountsList:
        writer.writerow([led['category_name']])
    return response


        #JOURNALS VIEWS

    
class JournalViewSet(viewsets.ModelViewSet):
    queryset = Journal.objects.all()
    serializer_class = JournalSerializer
    pagination_class = BasePagination

class JournalList(generics.ListCreateAPIView):
    queryset = Journal.objects.all()
    serializer_class = JournalSerializer
    pagination_class = DefaultPagination

class JournalDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = Journal.objects.all()
    serializer_class = JournalSerializer
        

@csrf_exempt
@api_view(['POST'])
def createJournal(request):
    company_id = request.data.get("company")
    txn_type = request.data.get("txn_type")
    done_by = request.data.get("done_by")

    if txn_type == 'JNL' and done_by is not None:
        done_by = request.user.first_name + ' '+ request.user.last_name
        company_uuid = uuid.UUID(company_id)
        last_journal_no = Journal.objects.filter(company=company_uuid, txn_type=txn_type).order_by('journal_no').last()

        if not last_journal_no:
            new_journal_no = "JNL-0001"
            company = get_object_or_404(Company, company_id=company_uuid)
            serializer = JournalSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(company=company, journal_no=new_journal_no, done_by=done_by)
            else:
                print(serializer.errors) 
            return Response(serializer.data)
        else:
            journal_no = last_journal_no.journal_no
            journal_no_int = int(journal_no.split('JNL-')[-1])
            new_journal_no_int = journal_no_int + 1
            new_journal_no = 'JNL-'+ str(new_journal_no_int).zfill(4)
            company = get_object_or_404(Company, company_id=company_uuid)
            serializer = JournalSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(company=company, journal_no=new_journal_no, done_by=done_by)
            else:
                print(serializer.errors) 
            return Response(serializer.data)
        
    
    elif txn_type == 'INV':
        company_uuid = uuid.UUID(company_id)
        last_journal_no = Journal.objects.filter(company=company_uuid, txn_type=txn_type).order_by('journal_no').last()

        if not last_journal_no:
            new_journal_no = "INV00001"
            company = get_object_or_404(Company, company_id=company_uuid)
            serializer = JournalSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(company=company, journal_no=new_journal_no)
            else:
                print(serializer.errors) 
            return Response(serializer.data)
        else:
            journal_no = last_journal_no.journal_no
            journal_no_int = int(journal_no.split('INV')[-1])
            new_journal_no_int = journal_no_int + 1
            new_journal_no = 'INV'+ str(new_journal_no_int).zfill(5)
            company = get_object_or_404(Company, company_id=company_uuid)
            serializer = JournalSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(company=company, journal_no=new_journal_no)
            else:
                print(serializer.errors) 
            return Response(serializer.data)
    
    elif txn_type == 'RCPT':
        company_uuid = uuid.UUID(company_id)
        last_journal_no = Journal.objects.filter(company=company_uuid, txn_type=txn_type).order_by('journal_no').last()

        if not last_journal_no:
            new_journal_no = "RC00001"
            company = get_object_or_404(Company, company_id=company_uuid)
            serializer = JournalSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(company=company, journal_no=new_journal_no, done_by=done_by)
            else:
                print(serializer.errors) 
            return Response(serializer.data)
        else:
            journal_no = last_journal_no.journal_no
            journal_no_int = int(journal_no.split('RC')[-1])
            new_journal_no_int = journal_no_int + 1
            new_journal_no = 'RC'+ str(new_journal_no_int).zfill(5)
            company = get_object_or_404(Company, company_id=company_uuid)
            serializer = JournalSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(company=company, journal_no=new_journal_no, done_by=done_by)
            else:
                print(serializer.errors) 
            return Response(serializer.data)
    
    elif txn_type == 'PMT':
        company_uuid = uuid.UUID(company_id)
        last_journal_no = Journal.objects.filter(company=company_uuid, txn_type=txn_type).order_by('journal_no').last()

        if not last_journal_no:
            new_journal_no = "PM00001"
            company = get_object_or_404(Company, company_id=company_uuid)
            serializer = JournalSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(company=company, journal_no=new_journal_no, done_by=done_by)
            else:
                print(serializer.errors) 
            return Response(serializer.data)
        else:
            journal_no = last_journal_no.journal_no
            journal_no_int = int(journal_no.split('PM')[-1])
            new_journal_no_int = journal_no_int + 1
            new_journal_no = 'PM'+ str(new_journal_no_int).zfill(5)
            company = get_object_or_404(Company, company_id=company_uuid)
            serializer = JournalSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(company=company, journal_no=new_journal_no, done_by=done_by)
            else:
                print(serializer.errors) 
            return Response(serializer.data)
    


@csrf_exempt
@api_view(['POST'])
def getJournals(request):
    journal_id = request.data.get("journal")
    company_id = request.data.get("company")
    txn_type = request.data.get("txn_type")
    journal_ledger = request.data.get("journal_ledger")

    if journal_id is not None:
        company_uuid = uuid.UUID(company_id)
        journal_uuid = uuid.UUID(journal_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        journal = Journal.objects.get(company=company, journal_id=journal_uuid)

        serializer = JournalSerializer(journal)
        return Response(serializer.data)
    
    elif txn_type is not None:
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        journals = Journal.objects.filter(company=company,txn_type= txn_type)

        serializer = JournalSerializer(journals, many=True)
        return Response(serializer.data)
    
    elif journal_ledger is not None:
        journal_ledger_uuid = uuid.UUID(journal_ledger)
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        ledger = get_object_or_404(Ledger, ledger_id=journal_ledger_uuid)
        journals = Journal.objects.filter(company=company,journal_ledger= ledger)

        serializer = JournalSerializer(journals, many=True)
        return Response(serializer.data)

    else:
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        journals = Journal.objects.filter(company=company)

        serializer = JournalSerializer(journals, many=True)
        return Response(serializer.data)


@csrf_exempt
@api_view(['PUT'])
def updateJournal(request):
    journal_id = request.data.get("journal")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    journal_uuid = uuid.UUID(journal_id)
    journal = Journal.objects.get(company=company_id,journal_id=journal_uuid)
    
    serializer = JournalSerializer(journal, data=request.data)

    if serializer.is_valid():
        serializer.save()

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def deleteJournal(request):
    journal_id = request.data.get("journal")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    journal_uuid = uuid.UUID(journal_id)
    journal = Journal.objects.get(company=company_id,journal_id=journal_uuid)

    journal.delete() 
    message = {'msg':"The journal has been succesfully deleted"} 
    return Response(message)


@csrf_exempt 
@api_view(['POST'])  
def generate_journals_pdf(request):
    journalList = []
    new_data = json.dumps(request.data)
    data = json.loads(new_data)
    journal_no = data['journal_no']
    description = data['description']
    date_from = data['date_from']
    date_to = data['date_to']
    min_amount = data['min_amount']
    max_amount = data['max_amount']
    txn_type = data['txn_type']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_journals = Journal.objects.filter(company=company_uuid)

    journals = company_journals.filter(Q(journal_no__icontains=journal_no) & Q(description__icontains=description) & Q(txn_type__icontains=txn_type))
    if date_from:
        journals = journals.filter(issue_date__gte=date_from)

    if date_to:
        journals = journals.filter(issue_date__lte=date_to)

    if min_amount:
        journals = journals.filter(total_amount__gte=min_amount)

    if max_amount:
        journals = journals.filter(total_amount__lte=max_amount)

    for jnl in journals:
        obj = {
            "journal_id": jnl.journal_id,
            "journal_no": jnl.journal_no,
            "client": jnl.client,
            "issue_date": jnl.issue_date.strftime("%d %b, %Y"),
            "due_date": jnl.due_date,
            "sub_total": jnl.sub_total,
            "tax": jnl.tax,
            "total_amount": jnl.total_amount,
            "total_paid": jnl.total_paid,
            "due_amount": jnl.due_amount,
            "status": jnl.status,
            "description": jnl.description,
            "done_by": jnl.done_by,

        }
        journalList.append(obj)

    context = {"journals":journalList}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/templates/financial_accounts_chart_of_accounts')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('journalsPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Journals.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/static/financial_accounts_chart_of_accounts/journalsPDF.css")

    path = 'Journals.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Journals.pdf'
    pdf.close()
    os.remove("Journals.pdf")  # remove the locally created pdf file.
    return response

@csrf_exempt
@api_view(['POST'])
def generate_journals_excel(request):
    journalList = []
    new_data = json.dumps(request.data)
    data = json.loads(new_data)
    journal_no = data['journal_no']
    description = data['description']
    date_from = data['date_from']
    date_to = data['date_to']
    min_amount = data['min_amount']
    max_amount = data['max_amount']
    txn_type = data['txn_type']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_journals = Journal.objects.filter(company=company_uuid)

    journals = company_journals.filter(Q(journal_no__icontains=journal_no) & Q(description__icontains=description) & Q(txn_type__icontains=txn_type))

    if date_from:
        journals = journals.filter(issue_date__gte=date_from)

    if date_to:
        journals = journals.filter(issue_date__lte=date_to)

    if min_amount:
        journals = journals.filter(total_amount__gte=min_amount)

    if max_amount:
        journals = journals.filter(total_amount__lte=max_amount)

    for jnl in journals:
        obj = {
            "journal_id": jnl.journal_id,
            "journal_no": jnl.journal_no,
            "client": jnl.client,
            "issue_date": jnl.issue_date.strftime("%d %b, %Y"),
            "due_date": jnl.due_date,
            "sub_total": jnl.sub_total,
            "tax": jnl.tax,
            "total_amount": jnl.total_amount,
            "total_paid": jnl.total_paid,
            "due_amount": jnl.due_amount,
            "status": jnl.status,
            "description": jnl.description,
            "done_by": jnl.done_by,

        }
        journalList.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Journals.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Journals")

    row_num = 0
    columns = ['Journal No', 'Date','Description','Amount','Done By']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style=style1)

    for jnl in journalList:
        row_num += 1
        row = [jnl['journal_no'],jnl['issue_date'],jnl['description'],jnl['total_amount'],jnl['done_by']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@csrf_exempt
@api_view(['POST'])
def generate_journals_csv(request):
    journalList = []
    new_data = json.dumps(request.data)
    data = json.loads(new_data)
    journal_no = data['journal_no']
    description = data['description']
    date_from = data['date_from']
    date_to = data['date_to']
    min_amount = data['min_amount']
    max_amount = data['max_amount']
    txn_type = data['txn_type']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_journals = Journal.objects.filter(company=company_uuid)

    journals = company_journals.filter(Q(journal_no__icontains=journal_no) & Q(description__icontains=description) & Q(txn_type__icontains=txn_type))

    if date_from:
        journals = journals.filter(issue_date__gte=date_from)

    if date_to:
        journals = journals.filter(issue_date__lte=date_to)

    if min_amount:
        journals = journals.filter(total_amount__gte=min_amount)

    if max_amount:
        journals = journals.filter(total_amount__lte=max_amount)

    for jnl in journals:
        obj = {
            "journal_id": jnl.journal_id,
            "journal_no": jnl.journal_no,
            "client": jnl.client,
            "issue_date": jnl.issue_date.strftime("%d %b, %Y"),
            "due_date": jnl.due_date,
            "sub_total": jnl.sub_total,
            "tax": jnl.tax,
            "total_amount": jnl.total_amount,
            "total_paid": jnl.total_paid,
            "due_amount": jnl.due_amount,
            "status": jnl.status,
            "description": jnl.description,
            "done_by": jnl.done_by,

        }
        journalList.append(obj)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Journals.csv'

    writer = csv.writer(response)
    writer.writerow(['Journal No', 'Date','Description','Amount','Done By'])

    for jnl in journalList:
        writer.writerow([jnl['journal_no'],jnl['issue_date'],jnl['description'],jnl['total_amount'],jnl['done_by']])
    return response


@csrf_exempt
@api_view(['POST'])
def patientInvoicePDF(request):
    invoice = request.data.get("invoice")
    hospital_id = request.data.get("hospital")
    patientInvoice = get_object_or_404(Journal, pk=invoice)
    hospital = Company.objects.get(company_id=hospital_id)

    invoice_no = patientInvoice.journal_no
    client = patientInvoice.client
    quantity = 1
    total_amount = patientInvoice.total_amount
    tax = patientInvoice.tax
    issue_date = patientInvoice.issue_date.strftime("%d %b, %Y")
    due_date = (datetime.strptime(issue_date, "%d %b, %Y")+ timedelta(days=3)).strftime("%d %b, %Y")
    sub_total = patientInvoice.sub_total
    description = patientInvoice.description


    context={"hospital": hospital,"invoice_no":invoice_no, "client":client, "tax":tax,
              "total_amount":total_amount, "due_date":due_date , "invoice_date":issue_date, "sub_total":sub_total,
              "description":description, "quantity":quantity}
    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/templates/financial_accounts_chart_of_accounts/')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('patientInvoicePDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Invoice.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/static/financial_accounts_chart_of_accounts/patientInvoicePDF.css")

    path = 'Invoice.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Invoice.pdf'
    pdf.close()
    os.remove("Invoice.pdf")  # remove the locally created pdf file.
    return response




        #JOURNAL ENTRIES VIEWS

    
class JournalEntryViewSet(viewsets.ModelViewSet):
    queryset = JournalEntry.objects.all()
    serializer_class = JournalEntrySerializer
    pagination_class = BasePagination

class JournalEntryList(generics.ListCreateAPIView):
    queryset = JournalEntry.objects.all()
    serializer_class = JournalEntrySerializer
    pagination_class = DefaultPagination

class JournalEntryDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = JournalEntry.objects.all()
    serializer_class = JournalEntrySerializer
        

@csrf_exempt
@api_view(['POST'])
def createJournalEntry(request):
    company_id = request.data.get("company")
    journal_id = request.data.get("journal")
    ledger_id = request.data.get("posting_account")      

    company_uuid = uuid.UUID(company_id)
    journal_uuid = uuid.UUID(journal_id)
    ledger_uuid = uuid.UUID(ledger_id)
    company = get_object_or_404(Company, company_id=company_uuid)
    journal = get_object_or_404(Journal, journal_id=journal_uuid)
    ledger = get_object_or_404(Ledger, ledger_id=ledger_uuid)
    serializer = JournalEntrySerializer(data=request.data)

    if serializer.is_valid():
        serializer.save(company=company, journal=journal, posting_account=ledger)
    else:
        print(serializer.errors)
    
    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def getJournalEntries(request):
    journal_entry_id = request.data.get("journal_entry")
    company_id = request.data.get("company")
    journal = request.data.get("journal")
    ledger = request.data.get("posting_account")

    if journal_entry_id is not None:
        company_uuid = uuid.UUID(company_id)
        journal_uuid = uuid.UUID(journal)
        journal_entry_uuid = uuid.UUID(journal_entry_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        journal = get_object_or_404(Journal, journal_id=journal_uuid)
        journal_entry = JournalEntry.objects.get(company=company, journal=journal, journal_entry_id=journal_entry_uuid)

        serializer = JournalEntrySerializer(journal_entry)
        return Response(serializer.data)
    
    elif journal is not None:
        journal_uuid = uuid.UUID(journal)
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        journal = get_object_or_404(Journal, journal_id=journal_uuid)
        journal_entries = JournalEntry.objects.filter(company=company,journal= journal)

        serializer = JournalEntrySerializer(journal_entries, many=True)
        return Response(serializer.data)

    elif ledger is not None:
        ledger_uuid = uuid.UUID(ledger)
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        ledger = get_object_or_404(Ledger, ledger_id=ledger_uuid)
        journal_entries = JournalEntry.objects.filter(company=company,posting_account= ledger)

        serializer = JournalEntrySerializer(journal_entries, many=True)
        return Response(serializer.data)
    
    else:
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        journal_entries = JournalEntry.objects.filter(company=company)

        serializer = JournalEntrySerializer(journal_entries, many=True)
        return Response(serializer.data)


@csrf_exempt
@api_view(['PUT'])
def updateJournalEntry(request):
    journal_entry_id = request.data.get("journal_entry")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    journal_entry_uuid = uuid.UUID(journal_entry_id)
    journal_entry = JournalEntry.objects.get(company=company_id,journal_entry_id=journal_entry_uuid)
    
    serializer = JournalEntrySerializer(journal_entry, data=request.data)

    if serializer.is_valid():
        serializer.save()

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def deleteJournalEntry(request):
    journal_entry_id = request.data.get("journal_entry")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    journal_entry_uuid = uuid.UUID(journal_entry_id)
    journal_entry = JournalEntry.objects.get(company=company_id,journal_entry_id=journal_entry_uuid)

    journal_entry.delete() 
    message = {'msg':"The journal_entry has been succesfully deleted"} 
    return Response(message)


@csrf_exempt   
def generate_journal_entries_pdf(request):
    chartOfAccountsList = []
    data = json.loads(request.body)
    ledger_code = data['ledger_code']
    ledger_name = data['ledger_name']
    financial_statement = data['financial_statement']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_ledgers = Ledger.objects.filter(company=company_uuid)

    ledgers = company_ledgers.filter(Q(ledger_name__icontains=ledger_name) & Q(ledger_code__icontains=ledger_code))

    if financial_statement:
        ledgers = ledgers.filter(financial_statement = financial_statement)

    for led in ledgers:
        obj = {
            "ledger_id": led.ledger_id,
            "ledger_code": led.ledger_code,
            "ledger_name": led.ledger_name,
            "ledger_type": led.ledger_type,
            "financial_statement": led.financial_statement,
            "balance": led.balance,

        }
        chartOfAccountsList.append(obj)

    context = {"ledgerd":chartOfAccountsList}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/templates/financial_accounts_chart_of_accounts')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('ledgerPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Chart Of Accounts.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/financial_accounts_chart_of_accounts/static/financial_accounts_chart_of_accounts/clientCategoryPDF.css")

    path = 'Chart Of Accounts.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Chart Of Accounts.pdf'
    pdf.close()
    os.remove("Chart Of Accounts.pdf")  # remove the locally created pdf file.
    return response

@csrf_exempt
def generate_journal_entries_excel(request):
    chartOfAccountsList = []
    data = json.loads(request.body)
    ledger_code = data['ledger_code']
    ledger_name = data['ledger_name']
    financial_statement = data['financial_statement']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_ledgers = Ledger.objects.filter(company=company_uuid)

    ledgers = company_ledgers.filter(Q(ledger_name__icontains=ledger_name) & Q(ledger_code__icontains=ledger_code))

    if financial_statement:
        ledgers = ledgers.filter(financial_statement = financial_statement)

    for led in ledgers:
        obj = {
            "ledger_id": led.ledger_id,
            "ledger_code": led.ledger_code,
            "ledger_name": led.ledger_name,
            "ledger_type": led.ledger_type,
            "financial_statement": led.financial_statement,
            "balance": led.balance,

        }
        chartOfAccountsList.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Chart Of Accounts.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Chart Of Accounts")

    row_num = 0
    columns = ['Client Category Name']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style=style1)

    for led in chartOfAccountsList:
        row_num += 1
        row = [led['category_name']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@csrf_exempt
def generate_journal_entries_csv(request):
    chartOfAccountsList = []
    data = json.loads(request.body)
    ledger_code = data['ledger_code']
    ledger_name = data['ledger_name']
    financial_statement = data['financial_statement']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_ledgers = Ledger.objects.filter(company=company_uuid)

    ledgers = company_ledgers.filter(Q(ledger_name__icontains=ledger_name) & Q(ledger_code__icontains=ledger_code))

    if financial_statement:
        ledgers = ledgers.filter(financial_statement = financial_statement)

    for led in ledgers:
        obj = {
            "ledger_id": led.ledger_id,
            "ledger_code": led.ledger_code,
            "ledger_name": led.ledger_name,
            "ledger_type": led.ledger_type,
            "financial_statement": led.financial_statement,
            "balance": led.balance,

        }
        chartOfAccountsList.append(obj)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Chart Of Accounts.csv'

    writer = csv.writer(response)
    writer.writerow([ 'Client Category Name'])

    for led in chartOfAccountsList:
        writer.writerow([led['category_name']])
    return response

