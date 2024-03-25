import os
from django.db import transaction
from django.db.models import F
from django.shortcuts import render,get_object_or_404
from django.http import HttpResponse, JsonResponse
from .models import *
from .patient_code_generator import patient_code_gen
from financial_accounts_chart_of_accounts.journal_no_generator import invoice_number_gen
from company.models import Company
from financial_accounts_chart_of_accounts.models import *
from users.models import *
from .serializers import *
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework import generics,status
from rest_framework.response import Response
from datetime import datetime
#Pagination
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
#Pdf 
import jinja2
import pdfkit
#Excel
import xlwt
#CSV
import csv

import json
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt

import uuid

#Data Import
from openpyxl import load_workbook

# Create your views here.

class BasePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 1000

class DefaultPagination(PageNumberPagination):
    page_size = 1000
    page_size_query_param = 'page_size'
    max_page_size = 1000


        #PATIENT VIEWS

class PatientViewSet(viewsets.ModelViewSet):
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer
    pagination_class = BasePagination

class PatientList(generics.ListCreateAPIView):
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer
    pagination_class = DefaultPagination

class PatientDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer


@csrf_exempt
@api_view(['POST'])
def createPatient(request):
    hospital_id = request.data.get("hospital")
    emergency_contact_id = request.data.get("emergency_contact_person")
    hospital_uuid = uuid.UUID(hospital_id)
    emergency_contact_uuid = uuid.UUID(emergency_contact_id)
    hospital = get_object_or_404(Company, company_id=hospital_uuid)
    emergency_contact = EmergencyContactPerson.objects.get(contact_person_id=emergency_contact_uuid)
    serializer = PatientSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save(hospital=hospital,emergency_contact_person=emergency_contact)

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)

@transaction.atomic
@csrf_exempt
@api_view(['POST'])
def create_patient_with_nextofkin_visit_and_charges(request):
    hospital_id = request.data.get('hospital')
    contact_first_name = request.data.get('contact_first_name')
    contact_last_name = request.data.get('contact_last_name')
    contact_email = request.data.get('contact_email')
    contact_phone_number = request.data.get('contact_phone_number')
    patient_name = request.data.get('patient')
    first_name = request.data.get('first_name')
    last_name = request.data.get('last_name')
    email = request.data.get('email')
    birth_date = request.data.get('birth_date')
    phone_number = request.data.get('phone_number')
    city = request.data.get('city')
    gender = request.data.get('gender')
    id_number = request.data.get('id_number')
    address = request.data.get('address')
    country = request.data.get('country')
    doctor = request.data.get('doctor')
    is_doctor = True
    staff = request.data.get('staff')
    visit_notes = request.data.get('visit_notes')
    company = request.data.get('company')
    client = request.data.get('client')
    description = request.data.get('description')
    txn_type = request.data.get('txn_type')
    issue_date = request.data.get('issue_date')
    total_amount = request.data.get('total_amount')
    journal_entry_array = request.data.get('journal_entry_array')
    print("The journal array is ",journal_entry_array)

    if journal_entry_array is not None:
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        contact_person = EmergencyContactPerson.objects.create(first_name=contact_first_name, last_name=contact_last_name, email=contact_email,
                                                               phone_number=contact_phone_number, patient=patient_name, hospital=hospital)
        print("The contact person is ",contact_person)
        patient_code = patient_code_gen(hospital_id)
        print("The patient code is ",patient_code)
        patient = Patient.objects.create(hospital=hospital, first_name=first_name, last_name=last_name, email=email, phone_number=phone_number, gender=gender, 
                                         id_number=id_number, birth_date=birth_date, city=city, address=address, country=country, emergency_contact_person=contact_person, patient_code=patient_code)
        print("The patient is ",patient)
        patient_ledger = Ledger.objects.get(ledger_code=patient.patient_code)
        print("The patient ledger is ",patient_ledger)

        if staff is not None and doctor is None:
            staff_uuid = uuid.UUID(staff)
            staff_to_visit = User.objects.get(allowed_company=hospital_uuid, user_id=staff_uuid)
            PatientHistory.objects.create(patient=patient, date=issue_date, notes=visit_notes, staff=staff_to_visit, hospital=hospital)
        elif doctor is not None and staff is None:
            staff_uuid = uuid.UUID(doctor)
            staff_to_visit = User.objects.get(allowed_company=hospital_uuid, user_id=staff_uuid)
            PatientHistory.objects.create(patient=patient, date=issue_date, notes=visit_notes, staff=staff_to_visit, is_doctor=is_doctor, hospital=hospital)

        invoice_number = invoice_number_gen(hospital_id)
        print("The invoice number is ",invoice_number)
        invoice_journal = Journal.objects.create(journal_no=invoice_number,txn_type=txn_type, client=client, description=description, issue_date=issue_date,
                                                 client_id=patient.patient_id, total_amount=total_amount, company=hospital)
        print("The invoice journal is ",invoice_journal)
        for jnlEntry in journal_entry_array:
            print("The JNLE is ", jnlEntry)
            print("The patient ledger is ",patient_ledger.ledger_id)
            if jnlEntry['posting_account'] == "":
                jnlEntry['posting_account'] = patient_ledger.ledger_id
                print("The posting account is ", jnlEntry['posting_account'])
                ledger_uuid = uuid.UUID(str(patient_ledger.ledger_id))
                company = get_object_or_404(Company, company_id=hospital_uuid)
                ledger = get_object_or_404(Ledger, ledger_id=ledger_uuid)
                JournalEntry.objects.create(journal=invoice_journal, date=jnlEntry['date'], description=jnlEntry['description'], txn_type=jnlEntry['txn_type'],
                                            posting_account=ledger, debit_amount=jnlEntry['debit_amount'], credit_amount=jnlEntry['credit_amount'], company=company)
               
            else:
                ledger_uuid = uuid.UUID(jnlEntry['posting_account'])
                company = get_object_or_404(Company, company_id=hospital_uuid)
                ledger = get_object_or_404(Ledger, ledger_id=ledger_uuid)
                JournalEntry.objects.create(journal=invoice_journal, date=jnlEntry['date'], description=jnlEntry['description'], txn_type=jnlEntry['txn_type'],
                                            posting_account=ledger, debit_amount=jnlEntry['debit_amount'], credit_amount=jnlEntry['credit_amount'], company=company)

        message = {'msg':"The patient, visit and charges have been succesfully added"}    
        return Response(message)

    elif staff is not None or doctor is not None and journal_entry_array is None:
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        contact_person = EmergencyContactPerson.objects.create(first_name=contact_first_name, last_name=contact_last_name, email=contact_email,
                                                               phone_number=contact_phone_number, patient=patient_name, hospital=hospital)
        print("The contact person is ",contact_person)
        patient_code = patient_code_gen(hospital_id)
        print("The patient code is ",patient_code)
        patient = Patient.objects.create(hospital=hospital, first_name=first_name, last_name=last_name, email=email, phone_number=phone_number, gender=gender, 
                                         id_number=id_number, birth_date=birth_date, city=city, address=address, country=country, emergency_contact_person=contact_person, patient_code=patient_code)
        print("The patient is ",patient)
        patient_ledger = Ledger.objects.get(ledger_code=patient.patient_code)
        print("The patient ledger is ",patient_ledger)

        if staff is not None and doctor is None:
            staff_uuid = uuid.UUID(staff)
            staff_to_visit = User.objects.get(allowed_company=hospital_uuid, user_id=staff_uuid)
            PatientHistory.objects.create(patient=patient, date=issue_date, notes=visit_notes, staff=staff_to_visit, hospital=hospital)
        elif doctor is not None and staff is None:
            staff_uuid = uuid.UUID(doctor)
            staff_to_visit = User.objects.get(allowed_company=hospital_uuid, user_id=staff_uuid)
            PatientHistory.objects.create(patient=patient, date=issue_date, notes=visit_notes, staff=staff_to_visit, is_doctor=is_doctor, hospital=hospital)

        message = {'msg':"The patient and visit without the charges have been succesfully added"}    
        return Response(message)

    else:
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        contact_person = EmergencyContactPerson.objects.create(first_name=contact_first_name, last_name=contact_last_name, email=contact_email,
                                                               phone_number=contact_phone_number, patient=patient_name, hospital=hospital)
        print("The contact person is ",contact_person)
        patient_code = patient_code_gen(hospital_id)
        print("The patient code is ",patient_code)
        patient = Patient.objects.create(hospital=hospital, first_name=first_name, last_name=last_name, email=email, phone_number=phone_number, gender=gender, 
                                         id_number=id_number, birth_date=birth_date, city=city, address=address, country=country, emergency_contact_person=contact_person, patient_code=patient_code)
        print("The patient is ",patient)
        message = {'msg':"Patient Added Succesfully"}    
        return Response(message)



@csrf_exempt
@api_view(['POST'])
def getPatients(request):
    hospital_id = request.data.get("hospital")
    patient_id = request.data.get("patient")

    if patient_id is not None:
        hospital_uuid = uuid.UUID(hospital_id)
        patient_uuid = uuid.UUID(patient_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        patient = Patient.objects.get(hospital=hospital, patient_id=patient_uuid)

        serializer = PatientSerializer(patient)
        return Response(serializer.data)

    else:
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        patients = Patient.objects.filter(hospital=hospital)

        serializer = PatientSerializer(patients, many=True)
        return Response(serializer.data)


@csrf_exempt
@api_view(['PUT'])
def updatePatient(request):
    patient_id = request.data.get("patient")
    hospital = request.data.get("hospital")
    emergency_contact_id = request.data.get("emergency_contact_person")
    hospital_uuid = uuid.UUID(hospital)
    emergency_contact_uuid = uuid.UUID(emergency_contact_id)
    hospital_id = get_object_or_404(Company, company_id=hospital_uuid)
    emergency_contact = EmergencyContactPerson.objects.get(contact_person_id=emergency_contact_uuid)
    patient_uuid = uuid.UUID(patient_id)
    patient = Patient.objects.get(hospital=hospital_id,patient_id=patient_uuid)
    
    serializer = PatientSerializer(patient, data=request.data)

    if serializer.is_valid():
        serializer.save(emergency_contact_person=emergency_contact)

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def deletePatient(request):
    patient_id = request.data.get("patient")
    hospital = request.data.get("hospital")
    hospital_uuid = uuid.UUID(hospital)
    hospital_id = get_object_or_404(Company, company_id=hospital_uuid)
    patient_uuid = uuid.UUID(patient_id)
    patient = Patient.objects.get(hospital=hospital_id,patient_id=patient_uuid)

    patient.delete() 
    message = {'msg':"The patient has been succesfully deleted"} 
    return Response(message)


@csrf_exempt   
@api_view(['POST'])
def generate_patients_pdf(request):
    patients = []
    new_data = json.dumps(request.data)
    data = json.loads(new_data)
    first_name = data['first_name']
    last_name = data['last_name']
    phone_number = data['phone_number']
    id_number = data['id_number']
    gender = data['gender']
    birth_date = data['birth_date']
    hospital_id = data['hospital_id']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_patients = Patient.objects.filter(hospital=hospital_uuid)

    patientList = hospital_patients.filter(Q(first_name__icontains=first_name) & Q(last_name__icontains=last_name) & Q(birth_date__icontains=birth_date)
                                        & Q(phone_number__icontains=phone_number) & Q(id_number__icontains=id_number))

    if gender:
        patientList = patientList.filter(gender=gender)

    for pat in patientList:
        if(pat.emergency_contact_person):
            obj = {
                "patient_id": pat.patient_id,
                "patient_code": pat.patient_code,
                "first_name": pat.first_name,
                "last_name": pat.last_name,
                "email": pat.email,
                "id_number": pat.id_number,
                "phone_number": pat.phone_number,
                "gender": pat.gender,
                "city": pat.city,
                "address": pat.address,
                "country": pat.country,
                "birth_date": pat.birth_date.strftime("%d %b, %Y"),
                "emergency_contact_person_id": pat.emergency_contact_person.contact_person_id,
                "emergency_contact_person_name": pat.emergency_contact_person.first_name + " "+pat.emergency_contact_person.last_name,
                "emergency_contact_person_email": pat.emergency_contact_person.email,
                "emergency_contact_person_phone_number": pat.emergency_contact_person.phone_number,
            }
            patients.append(obj)
        else:
            obj = {
                "patient_id": pat.patient_id,
                "patient_code": pat.patient_code,
                "first_name": pat.first_name,
                "last_name": pat.last_name,
                "email": pat.email,
                "id_number": pat.id_number,
                "phone_number": pat.phone_number,
                "gender": pat.gender,
                "city": pat.city,
                "address": pat.address,
                "country": pat.country,
                "birth_date": pat.birth_date.strftime("%d %b, %Y"),
            }
            patients.append(obj)

    context = {"patients":patients}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/patients_registration/templates/patients_registration')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('patientsPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Patients.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/patients_registration/static/patients_registration/patientsPDF.css")

    path = 'Patients.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Patients.pdf'
    pdf.close()
    os.remove("Patients.pdf")  # remove the locally created pdf file.
    return response

@csrf_exempt
@api_view(['POST'])
def generate_patients_excel(request):
    patients = []
    empty = ""
    new_data = json.dumps(request.data)
    data = json.loads(new_data)
    first_name = data['first_name']
    last_name = data['last_name']
    birth_date = data['birth_date']
    phone_number = data['phone_number']
    id_number = data['id_number']
    gender = data['gender']
    hospital_id = data['hospital_id']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_patients = Patient.objects.filter(hospital=hospital_uuid)

    patientList = hospital_patients.filter(Q(first_name__icontains=first_name) & Q(last_name__icontains=last_name) & Q(birth_date__icontains=birth_date)
                                        & Q(phone_number__icontains=phone_number) & Q(id_number__icontains=id_number))

    if gender:
        patientList = patientList.filter(gender=gender)

    for pat in patientList:
        if(pat.emergency_contact_person):
            obj = {
                "patient_id": pat.patient_id,
                "patient_code": pat.patient_code,
                "first_name": pat.first_name,
                "last_name": pat.last_name,
                "email": pat.email,
                "id_number": pat.id_number,
                "phone_number": pat.phone_number,
                "gender": pat.gender,
                "city": pat.city,
                "address": pat.address,
                "country": pat.country,
                "birth_date": pat.birth_date.strftime("%d %b, %Y"),
                "emergency_contact_person_id": pat.emergency_contact_person.contact_person_id,
                "emergency_contact_person_name": pat.emergency_contact_person.first_name + " "+pat.emergency_contact_person.last_name,
                "emergency_contact_person_email": pat.emergency_contact_person.email,
                "emergency_contact_person_phone_number": pat.emergency_contact_person.phone_number,
            }
            patients.append(obj)
        else:
            obj = {
                "patient_id": pat.patient_id,
                "patient_code": pat.patient_code,
                "first_name": pat.first_name,
                "last_name": pat.last_name,
                "email": pat.email,
                "id_number": pat.id_number,
                "phone_number": pat.phone_number,
                "gender": pat.gender,
                "city": pat.city,
                "address": pat.address,
                "country": pat.country,
                "birth_date": pat.birth_date.strftime("%d %b, %Y"),
                "emergency_contact_person_id": empty,
                "emergency_contact_person_name": empty,
                "emergency_contact_person_email": empty,
                "emergency_contact_person_phone_number": empty,
            }
            patients.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Patients.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Patients")

    row_num = 0
    columns = ['Code','First Name','Last Name','Email','ID Number','Phone Number','Gender','Address','Birth Date','City','Country','Contact Person','Phone Number','Email']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style=style1)

    for pat in patients:
        row_num += 1
        row = [pat['patient_code'],pat['first_name'],pat['last_name'],pat['email'],pat['id_number'],pat['phone_number'],pat['gender'],pat['address'],pat['birth_date'],pat['city'],pat['country'],pat['emergency_contact_person_name'],pat['emergency_contact_person_phone_number'],pat['emergency_contact_person_email']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@csrf_exempt
@api_view(['POST'])
def generate_patients_csv(request):
    patients = []
    empty = ""
    new_data = json.dumps(request.data)
    data = json.loads(new_data)
    first_name = data['first_name']
    last_name = data['last_name']
    birth_date = data['birth_date']
    phone_number = data['phone_number']
    id_number = data['id_number']
    gender = data['gender']
    hospital_id = data['hospital_id']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_patients = Patient.objects.filter(hospital=hospital_uuid)

    patientList = hospital_patients.filter(Q(first_name__icontains=first_name) & Q(last_name__icontains=last_name) & Q(birth_date__icontains=birth_date)
                                        & Q(phone_number__icontains=phone_number) & Q(id_number__icontains=id_number))

    if gender:
        patientList = patientList.filter(gender=gender)

    for pat in patientList:
        if(pat.emergency_contact_person):
            obj = {
                "patient_id": pat.patient_id,
                "patient_code": pat.patient_code,
                "first_name": pat.first_name,
                "last_name": pat.last_name,
                "email": pat.email,
                "id_number": pat.id_number,
                "phone_number": pat.phone_number,
                "gender": pat.gender,
                "city": pat.city,
                "address": pat.address,
                "country": pat.country,
                "birth_date": pat.birth_date.strftime("%d %b, %Y"),
                "emergency_contact_person_id": pat.emergency_contact_person.contact_person_id,
                "emergency_contact_person_name": pat.emergency_contact_person.first_name + " "+pat.emergency_contact_person.last_name,
                "emergency_contact_person_email": pat.emergency_contact_person.email,
                "emergency_contact_person_phone_number": pat.emergency_contact_person.phone_number,
            }
            patients.append(obj)
        else:
            obj = {
                "patient_id": pat.patient_id,
                "patient_code": pat.patient_code,
                "first_name": pat.first_name,
                "last_name": pat.last_name,
                "email": pat.email,
                "id_number": pat.id_number,
                "phone_number": pat.phone_number,
                "gender": pat.gender,
                "city": pat.city,
                "address": pat.address,
                "country": pat.country,
                "birth_date": pat.birth_date.strftime("%d %b, %Y"),
                "emergency_contact_person_id": empty,
                "emergency_contact_person_name": empty,
                "emergency_contact_person_email": empty,
                "emergency_contact_person_phone_number": empty,
            }
            patients.append(obj)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Patients.csv'

    writer = csv.writer(response)
    writer.writerow(['Code','First Name','Last Name','Email','ID Number','Phone Number','Gender','Address','Birth Date','City','Country','Contact Person','Phone Number','Email'])

    for pat in patients:
        writer.writerow([pat['patient_code'],pat['first_name'],pat['last_name'],pat['email'],pat['id_number'],pat['phone_number'],pat['gender'],pat['address'],pat['birth_date'],pat['city'],pat['country'],pat['emergency_contact_person_name'],pat['emergency_contact_person_phone_number'],pat['emergency_contact_person_email']])
    return response

@csrf_exempt
@api_view(['POST'])
def display_patients_import_excel(request):
    patientList = []
    excel_file = request.FILES['patients_excel']
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_name,last_name,email,id_number,phone_number,gender,address,birth_date,city,country, contact_first_name,contact_last_name,contact_number,contact_email = row
        obj = {
            "first_name": first_name,
            "last_name": last_name,
            "email": email,
            "id_number": id_number,
            "phone_number": phone_number,
            "gender": gender,
            "address": address,
            "birth_date": datetime.strptime(str(birth_date), "%Y-%m-%d %H:%M:%S").strftime("%d %b, %Y"),
            "city": city,
            "country": country,
        }
        patientList.append(obj)

    return JsonResponse({"patients": patientList})


@csrf_exempt
@api_view(['POST'])
def import_patients_excel(request):

    excel_file = request.FILES['patients_excel']
    hospital_id = request.data.get('hospital_id')
    hospital_uuid = uuid.UUID(hospital_id)
    hospital = get_object_or_404(Company, company_id=hospital_uuid)
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_name,last_name,email,id_number,phone_number,gender,address,birth_date,city,country, contact_first_name,contact_last_name,contact_number,contact_email = row
        pat_code = patient_code_gen(hospital_id)
        patient_name = first_name + " "+ last_name  
        patient = Patient.objects.create(patient_code=pat_code, first_name=first_name, last_name=last_name, email=email, id_number=id_number, birth_date=birth_date,
                               phone_number=phone_number,city=city, gender=gender, address=address, country=country, hospital=hospital) 
        
        if(contact_first_name or contact_last_name):
            contact_person = EmergencyContactPerson.objects.create(first_name=contact_first_name,last_name=contact_last_name,email=contact_email, phone_number=contact_number, patient=patient_name, hospital=hospital)
            contact_person.save()
            Patient.objects.filter(patient_code=pat_code).update(emergency_contact_person=contact_person)
        else:
            patient.save()
        

    return HttpResponse("Excel Import Succesful")


        #EMERGENCY CONTACT PERSON VIEWS

class EmergencyContactPersonViewSet(viewsets.ModelViewSet):
    queryset = EmergencyContactPerson.objects.all()
    serializer_class = EmergencyContactPersonSerializer
    pagination_class = BasePagination

class EmergencyContactPersonList(generics.ListCreateAPIView):
    queryset = EmergencyContactPerson.objects.all()
    serializer_class = EmergencyContactPersonSerializer
    pagination_class = DefaultPagination

class EmergencyContactPersonDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = EmergencyContactPerson.objects.all()
    serializer_class = EmergencyContactPersonSerializer


@csrf_exempt
@api_view(['POST'])
def createEmergencyContactPerson(request):
    hospital_id = request.data.get("hospital")
    hospital_uuid = uuid.UUID(hospital_id)
    hospital = get_object_or_404(Company, company_id=hospital_uuid)
    serializer = EmergencyContactPersonSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save(hospital=hospital)

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def getEmergencyContactPersons(request):
    hospital_id = request.data.get("hospital")
    contact_person_id = request.data.get("contact_person")

    if contact_person_id is not None:
        hospital_uuid = uuid.UUID(hospital_id)
        contact_person_uuid = uuid.UUID(contact_person_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        contact_person = EmergencyContactPerson.objects.get(hospital=hospital, contact_person_id=contact_person_uuid)

        serializer = EmergencyContactPersonSerializer(contact_person)
        return Response(serializer.data)

    else:
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        contact_persons = EmergencyContactPerson.objects.filter(hospital=hospital)

        serializer = EmergencyContactPersonSerializer(contact_persons, many=True)
        return Response(serializer.data)


@csrf_exempt
@api_view(['PUT'])
def updateEmergencyContactPerson(request):
    contact_person_id = request.data.get("contact_person")
    hospital = request.data.get("hospital")
    hospital_uuid = uuid.UUID(hospital)
    hospital_id = get_object_or_404(Company, company_id=hospital_uuid)
    contact_person_uuid = uuid.UUID(contact_person_id)
    contact_person = EmergencyContactPerson.objects.get(hospital=hospital_id,contact_person_id=contact_person_uuid)
    
    serializer = EmergencyContactPersonSerializer(contact_person, data=request.data)

    if serializer.is_valid():
        serializer.save()

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def deleteEmergencyContactPerson(request):
    contact_person_id = request.data.get("contact_person")
    hospital = request.data.get("hospital")
    hospital_uuid = uuid.UUID(hospital)
    hospital_id = get_object_or_404(Company, company_id=hospital_uuid)
    contact_person_uuid = uuid.UUID(contact_person_id)
    contact_person = EmergencyContactPerson.objects.get(hospital=hospital_id,contact_person_id=contact_person_uuid)

    contact_person.delete() 
    message = {'msg':"The emergency contact person has been succesfully deleted"} 
    return Response(message)


@csrf_exempt   
def generate_contact_persons_pdf(request):
    contact_persons = []
    data = json.loads(request.body)
    first_name = data['first_name']
    last_name = data['last_name']
    email = data['email']
    phone_number = data['phone_number']
    hospital_id = data['hospital_id']

    hospital_uuid = uuid.UUID(hospital_id)
    patients_contact_persons = EmergencyContactPerson.objects.filter(hospital=hospital_uuid)

    contactPersonList = patients_contact_persons.filter(Q(first_name__icontains=first_name) & Q(last_name__icontains=last_name)
                                                       & Q(email__icontains=email) & Q(phone_number__icontains=phone_number))
                                        

    for cont in contactPersonList:
        obj = {
            "contact_person_id": cont.contact_person_id,
            "first_name": cont.first_name,
            "last_name": cont.last_name,
            "email": cont.email,
            "phone_number": cont.phone_number,
            "patient": cont.patient
        }
        contact_persons.append(obj)

    context = {"contact_persons":contact_persons}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/patients_registration/templates/patients_registration')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('contactPersonsPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Contact People.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/patients_registration/static/patients_registration/patientsPDF.css")

    path = 'Contact People.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Contact People.pdf'
    pdf.close()
    os.remove("Contact People.pdf")  # remove the locally created pdf file.
    return response

@csrf_exempt
def generate_contact_persons_excel(request):
    contact_persons = []
    data = json.loads(request.body)
    first_name = data['first_name']
    last_name = data['last_name']
    email = data['email']
    phone_number = data['phone_number']
    hospital_id = data['hospital_id']

    hospital_uuid = uuid.UUID(hospital_id)
    patients_contact_persons = EmergencyContactPerson.objects.filter(hospital=hospital_uuid)

    contactPersonList = patients_contact_persons.filter(Q(first_name__icontains=first_name) & Q(last_name__icontains=last_name)
                                                       & Q(email__icontains=email) & Q(phone_number__icontains=phone_number))
                                        

    for cont in contactPersonList:
        obj = {
            "contact_person_id": cont.contact_person_id,
            "first_name": cont.first_name,
            "last_name": cont.last_name,
            "email": cont.email,
            "phone_number": cont.phone_number,
            "patient": cont.patient
        }
        contact_persons.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Contact People.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Contact People")

    row_num = 0
    columns = ['First Name','Last Name','Email','Phone Number','Patient']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style=style1)

    for cont in contact_persons:
        row_num += 1
        row = [cont['first_name'],cont['last_name'],cont['email'],cont['phone_number'],cont['patient']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@csrf_exempt
def generate_contact_persons_csv(request):
    contact_persons = []
    data = json.loads(request.body)
    first_name = data['first_name']
    last_name = data['last_name']
    email = data['email']
    phone_number = data['phone_number']
    hospital_id = data['hospital_id']

    hospital_uuid = uuid.UUID(hospital_id)
    patients_contact_persons = EmergencyContactPerson.objects.filter(hospital=hospital_uuid)

    contactPersonList = patients_contact_persons.filter(Q(first_name__icontains=first_name) & Q(last_name__icontains=last_name)
                                                       & Q(email__icontains=email) & Q(phone_number__icontains=phone_number))
                                        

    for cont in contactPersonList:
        obj = {
            "contact_person_id": cont.contact_person_id,
            "first_name": cont.first_name,
            "last_name": cont.last_name,
            "email": cont.email,
            "phone_number": cont.phone_number,
            "patient": cont.patient
        }
        contact_persons.append(obj)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Conatact People.csv'

    writer = csv.writer(response)
    writer.writerow(['First Name','Last Name','Email','Phone Number','Patient'])

    for cont in contact_persons:
        writer.writerow([cont['first_name'],cont['last_name'],cont['email'],cont['phone_number'],cont['patient']])
    return response


        #PATIENT HISTORY VIEWS

class PatientHistoryViewSet(viewsets.ModelViewSet):
    queryset = PatientHistory.objects.all()
    serializer_class = PatientHistorySerializer
    pagination_class = BasePagination

class PatientHistoryList(generics.ListCreateAPIView):
    queryset = PatientHistory.objects.all()
    serializer_class = PatientHistorySerializer
    pagination_class = DefaultPagination

class PatientHistoryDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = PatientHistory.objects.all()
    serializer_class = PatientHistorySerializer


@csrf_exempt
@api_view(['POST'])
def createPatientHistory(request):
    hospital_id = request.data.get("hospital")
    patient_id = request.data.get("patient")
    hospital_uuid = uuid.UUID(hospital_id)
    patient_uuid = uuid.UUID(patient_id)
    hospital = get_object_or_404(Company, company_id=hospital_uuid)
    patient = Patient.objects.get(patient_id=patient_uuid)
    serializer = PatientHistorySerializer(data=request.data)

    if serializer.is_valid():
        serializer.save(hospital=hospital,patient=patient)

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def getPatientHistories(request):
    hospital_id = request.data.get("hospital")
    patient_history_id = request.data.get("patient_history")

    if patient_history_id is not None:
        hospital_uuid = uuid.UUID(hospital_id)
        patient_history_uuid = uuid.UUID(patient_history_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        patient_history = PatientHistory.objects.get(hospital=hospital, patient_history_id=patient_history_uuid)

        serializer = PatientHistorySerializer(patient_history)
        return Response(serializer.data)

    else:
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        patient_history = PatientHistory.objects.filter(hospital=hospital)

        serializer = PatientHistorySerializer(patient_history, many=True)
        return Response(serializer.data)


@csrf_exempt
@api_view(['PUT'])
def updatePatientHistory(request):
    patient_history_id = request.data.get("patient_history")
    hospital = request.data.get("hospital")
    staff_id = request.data.get("staff")
    
    if staff_id is not None:
        hospital_uuid = uuid.UUID(hospital)
        staff_uuid = uuid.UUID(staff_id)
        patient_history_uuid = uuid.UUID(patient_history_id)
        hospital_id = get_object_or_404(Company, company_id=hospital_uuid)
        staff = User.objects.get(user_id=staff_uuid, allowed_company=hospital_id)
        patient_history = PatientHistory.objects.get(hospital=hospital_id,patient_history_id=patient_history_uuid)
        serializer = PatientHistorySerializer(patient_history, data=request.data)

        if serializer.is_valid():
            serializer.save(staff=staff)
        else:
            print(serializer.errors) 
        return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def deletePatientHistory(request):
    patient_history_id = request.data.get("patient_history")
    hospital = request.data.get("hospital")
    hospital_uuid = uuid.UUID(hospital)
    hospital_id = get_object_or_404(Company, company_id=hospital_uuid)
    patient_history_uuid = uuid.UUID(patient_history_id)
    patient_history = PatientHistory.objects.get(hospital=hospital_id,patient_history_id=patient_history_uuid)

    patient_history.delete() 
    message = {'msg':"The patient_history has been succesfully deleted"} 
    return Response(message)


@csrf_exempt   
@api_view(['POST'])
def generate_patients_history_pdf(request):
    patientHistoryList = []
    new_data = json.dumps(request.data)
    data = json.loads(new_data)
    patient = data['patient']
    date_from = data['date_from']
    date_to = data['date_to']
    staff = data['staff']
    patient_code = data['patient_code']
    hospital_id = data['hospital_id']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_patients_history = PatientHistory.objects.filter(hospital=hospital_uuid)

    patients_history = hospital_patients_history.filter((Q(patient__first_name__icontains=patient) | Q(patient__last_name__icontains=patient))
                                         & Q(patient__patient_code__icontains=patient_code))
    
    if date_from:
        patients_history = patients_history.filter(date__gte=date_from) 
    
    if date_to:
        patients_history = patients_history.filter(date__lte=date_to) 

    if staff:
        patients_history = patients_history.filter((Q(staff__first_name__icontains=staff) | Q(staff__last_name__icontains=staff)))


    for hist in patients_history:
        obj = {
            "patient_history_id": hist.patient_history_id,
            "patient_code": hist.patient.patient_code,
            "patient_name": hist.patient.first_name+ ' '+hist.patient.last_name,
            "patient_id_number": hist.patient.id_number,
            "patient_id": hist.patient.patient_id,
            "date": hist.date.strftime("%d %b, %Y"),
            "notes": hist.notes,
            "staff_name": hist.staff.first_name+ ' '+hist.staff.last_name,
            "staff_id": hist.staff.user_id,
            "staff_profile": hist.staff.profile,
            "staff_is_doctor": hist.is_doctor
            
        }
        patientHistoryList.append(obj)

    context = {"patients_visits":patientHistoryList}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/patients_registration/templates/patients_registration')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('patientsHistoryPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Patients Visits.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/patients_registration/static/patients_registration/patientsPDF.css")

    path = 'Patients Visits.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Patients Visits.pdf'
    pdf.close()
    os.remove("Patients Visits.pdf")  # remove the locally created pdf file.
    return response

@csrf_exempt
@api_view(['POST'])
def generate_patients_history_excel(request):
    patientHistoryList = []
    new_data = json.dumps(request.data)
    data = json.loads(new_data)
    patient = data['patient']
    date_from = data['date_from']
    date_to = data['date_to']
    staff = data['staff']
    patient_code = data['patient_code']
    hospital_id = data['hospital_id']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_patients_history = PatientHistory.objects.filter(hospital=hospital_uuid)

    patients_history = hospital_patients_history.filter((Q(patient__first_name__icontains=patient) | Q(patient__last_name__icontains=patient))
                                         & Q(patient__patient_code__icontains=patient_code))
    
    if date_from:
        patients_history = patients_history.filter(date__gte=date_from) 
    
    if date_to:
        patients_history = patients_history.filter(date__lte=date_to) 

    if staff:
        patients_history = patients_history.filter((Q(staff__first_name__icontains=staff) | Q(staff__last_name__icontains=staff)))


    for hist in patients_history:
        obj = {
            "patient_history_id": hist.patient_history_id,
            "patient_code": hist.patient.patient_code,
            "patient_name": hist.patient.first_name+ ' '+hist.patient.last_name,
            "patient_id_number": hist.patient.id_number,
            "patient_id": hist.patient.patient_id,
            "date": hist.date.strftime("%d %b, %Y"),
            "notes": hist.notes,
            "staff_name": hist.staff.first_name+ ' '+hist.staff.last_name,
            "staff_id": hist.staff.user_id,
            "staff_profile": hist.staff.profile,
            "staff_is_doctor": hist.is_doctor
            
        }
        patientHistoryList.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Patients Visits.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Patients Visits")

    row_num = 0
    columns = ['Date','Patient Code','Patient Name','Patient ID No','Staff Name','Notes']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style=style1)

    for hist in patientHistoryList:
        row_num += 1
        row = [hist['date'],hist['patient_code'],hist['patient_name'],hist['patient_id_number'],hist['staff_name'],hist['notes']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@csrf_exempt
@api_view(['POST'])
def generate_patients_history_csv(request):
    patientHistoryList = []
    new_data = json.dumps(request.data)
    data = json.loads(new_data)
    patient = data['patient']
    date_from = data['date_from']
    date_to = data['date_to']
    staff = data['staff']
    patient_code = data['patient_code']
    hospital_id = data['hospital_id']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_patients_history = PatientHistory.objects.filter(hospital=hospital_uuid)

    patients_history = hospital_patients_history.filter((Q(patient__first_name__icontains=patient) | Q(patient__last_name__icontains=patient))
                                         & Q(patient__patient_code__icontains=patient_code))
    
    if date_from:
        patients_history = patients_history.filter(date__gte=date_from) 
    
    if date_to:
        patients_history = patients_history.filter(date__lte=date_to) 

    if staff:
        patients_history = patients_history.filter((Q(staff__first_name__icontains=staff) | Q(staff__last_name__icontains=staff)))


    for hist in patients_history:
        obj = {
            "patient_history_id": hist.patient_history_id,
            "patient_code": hist.patient.patient_code,
            "patient_name": hist.patient.first_name+ ' '+hist.patient.last_name,
            "patient_id_number": hist.patient.id_number,
            "patient_id": hist.patient.patient_id,
            "date": hist.date.strftime("%d %b, %Y"),
            "notes": hist.notes,
            "staff_name": hist.staff.first_name+ ' '+hist.staff.last_name,
            "staff_id": hist.staff.user_id,
            "staff_profile": hist.staff.profile,
            "staff_is_doctor": hist.is_doctor
            
        }
        patientHistoryList.append(obj)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Patients Visits.csv'

    writer = csv.writer(response)
    writer.writerow(['Date','Patient Code','Patient Name','Patient ID No','Staff Name','Notes'])

    for hist in patientHistoryList:
        writer.writerow([hist['date'],hist['patient_code'],hist['patient_name'],hist['patient_id_number'],hist['staff_name'],hist['notes']])
    return response