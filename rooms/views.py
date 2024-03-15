import environ,os
# Initialise environment variables
env = environ.Env()
environ.Env.read_env()

import re
import json
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse, Http404
from .models import *
from company.models import Company
from xtra.models import Department
from .serializers import *
from rest_framework.views import APIView
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.decorators import api_view
from datetime import datetime,timedelta
#Pagination
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
#Pdf 
import jinja2
import pdfkit
#Excel
import xlwt
#CSV
import csv

import json
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt

import uuid

#Data Import
from openpyxl import load_workbook

# Create your views here.

        #PAGINATION

class BasePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 1000

class DefaultPagination(PageNumberPagination):
    page_size = 1000
    page_size_query_param = 'page_size'
    max_page_size = 1000


           #ROOMS VIEWS
    
class RoomList(generics.ListCreateAPIView):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer

class RoomDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer


@csrf_exempt
@api_view(['POST'])
def createRoom(request):
    company_id = request.data.get("company")
    department_id = request.data.get("department")
    company_uuid = uuid.UUID(company_id)
    department_uuid = uuid.UUID(department_id)
    company = get_object_or_404(Company, company_id=company_uuid)
    department = Department.objects.get(department_id=department_uuid)
    serializer = RoomSerializer(data=request.data)


    if serializer.is_valid():
        serializer.save(department=department,company=company)

    else:
        print(serializer.errors) 
    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def getDepartmentRooms(request):
    room_id = request.data.get("room")
    company_id = request.data.get("company")
    department_id = request.data.get("department")

    if room_id is not None and department_id is not None:
        department_uuid = uuid.UUID(department_id)
        company_uuid = uuid.UUID(company_id)
        room_uuid = uuid.UUID(room_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        room = Room.objects.get(company=company, department=department_uuid, room_id=room_uuid)

        serializer = RoomSerializer(room)
        return Response(serializer.data)

    elif department_id is not None:
        department_uuid = uuid.UUID(department_id)
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        room = Room.objects.filter(company=company, department=department_uuid)

        serializer = RoomSerializer(room, many=True)
        return Response(serializer.data)

    else:
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        rooms = Room.objects.filter(company=company)

        serializer = RoomSerializer(rooms, many=True)
        return Response(serializer.data)

csrf_exempt
@api_view(['PUT'])
def updateDepartmentRoom(request):
    room_id = request.data.get("room")
    department_id = request.data.get("department")
    company = request.data.get("company")
    room_uuid = uuid.UUID(room_id)
    department_uuid = uuid.UUID(department_id)
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    room = Room.objects.get(company=company_id,room_id=room_uuid)
    
    serializer = RoomSerializer(room, data=request.data)

    if serializer.is_valid():
        serializer.save()

    else:
        print(serializer.errors)

    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def deleteRoom(request):
    room_id = request.data.get("room")
    department_id = request.data.get("department")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    department_uuid = uuid.UUID(department_id)
    room_uuid = uuid.UUID(room_id)
    department = get_object_or_404(Department, company=company_uuid, department_id=department_uuid)
    room = Room.objects.get(company=company_uuid,department=department, room_id=room_uuid)

    room.delete() 
    message = {'msg':"The room has been succesfully deleted"} 
    return Response(message)

        #ROOM REPORTS

@csrf_exempt 
@api_view(['POST'])
def generate_rooms_pdf(request):
    roomsList = []
    data = json.loads(request.body)
    room_code = data['room_code']
    room_name = data['room_name']
    department = data['department']
    company_id = data['company']

    company_uuid = uuid.UUID(company_id)
    department_rooms = Room.objects.filter(company=company_uuid)

    rooms = department_rooms.filter(Q(room_code__icontains=room_code) & Q(room_name__icontains=room_name) & Q(department__name__icontains=department) )

    for rom in rooms:
        obj = {
            "room_id": rom.room_id,
            "room_code": rom.room_code,
            "room_name": rom.room_name,
            "department_id": rom.department.department_id,
            "department_name": rom.department.name,
            "staff": rom.staff
        }
        roomsList.append(obj)

    context = {"rooms":roomsList}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/rooms/templates/rooms')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('roomsPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Rooms.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/rooms/static/rooms/roomsPDF.css")

    path = 'Rooms.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Rooms.pdf'
    pdf.close()
    os.remove("Rooms.pdf")  # remove the locally created pdf file.
    return response

@api_view(['POST'])
@csrf_exempt
def generate_rooms_excel(request):
    roomsList = []
    data = json.loads(request.body)
    room_code = data['room_code']
    room_name = data['room_name']
    department = data['department']
    company_id = data['company']

    company_uuid = uuid.UUID(company_id)
    department_rooms = Room.objects.filter(company=company_uuid)

    rooms = department_rooms.filter(Q(room_code__icontains=room_code) & Q(room_name__icontains=room_name) & Q(department__name__icontains=department) )

    for rom in rooms:
        obj = {
            "room_id": rom.room_id,
            "room_code": rom.room_code,
            "room_name": rom.room_name,
            "department_id": rom.department.department_id,
            "department_name": rom.department.name,
            "staff": rom.staff
        }
        roomsList.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Rooms.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Rooms")

    row_num = 0
    columns = ['Code','Name','Department','Staff']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style = style1)

    for rom in roomsList:
        row_num += 1
        row = [rom['room_code'],rom['room_name'], rom['department_name'], rom['staff']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@api_view(['POST'])
@csrf_exempt
def generate_rooms_csv(request):
    roomsList = []
    data = json.loads(request.body)
    room_code = data['room_code']
    room_name = data['room_name']
    department = data['department']
    company_id = data['company']

    company_uuid = uuid.UUID(company_id)
    department_rooms = Room.objects.filter(company=company_uuid)

    rooms = department_rooms.filter(Q(room_code__icontains=room_code) & Q(room_name__icontains=room_name) & Q(department__name__icontains=department) )

    for rom in rooms:
        obj = {
            "room_id": rom.room_id,
            "room_code": rom.room_code,
            "room_name": rom.room_name,
            "department_id": rom.department.department_id,
            "department_name": rom.department.name,
            "staff": rom.staff
        }
        roomsList.append(obj)


    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Rooms.csv'

    writer = csv.writer(response)
    writer.writerow(['Code','Name','Department','Staff'])

    for rom in roomsList:
        writer.writerow([rom['room_code'],rom['room_name'], rom['department_name'], rom['staff']])
    return response

@csrf_exempt
@api_view(['POST'])
def display_rooms_import_excel(request):
    roomsList = []
    excel_file = request.FILES['rooms_excel']
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        code,name, department,staff = row
        obj = {
            "room_code": code,
            "room_name": name,
            "department": department,
            "staff": staff
        }
        roomsList.append(obj)

    return JsonResponse({"rooms": roomsList})


@csrf_exempt
@api_view(['POST'])
def import_rooms_excel(request):

    excel_file = request.FILES['rooms_excel']
    company_id = request.data.get('company')
    company_uuid = uuid.UUID(company_id)
    company = get_object_or_404(Company, company_id=company_uuid)
    rooms = Room.objects.filter(company=company)
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        code,name, department,staff = row
        department_id = Department.objects.get(name=department,company=company)
        Room.objects.create(room_code=code, room_name=name, department=department_id, company=company, staff=staff) 
        

    return HttpResponse("Excel Import Succesful")


          #WARDS VIEWS
    
class WardList(generics.ListCreateAPIView):
    queryset = Ward.objects.all()
    serializer_class = WardSerializer

class WardDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = Ward.objects.all()
    serializer_class = WardSerializer


@csrf_exempt
@api_view(['POST'])
def createWard(request):
    hospital_id = request.data.get("hospital")
    hospital_uuid = uuid.UUID(hospital_id)
    hospital = get_object_or_404(Company, company_id=hospital_uuid)

    serializer = WardSerializer(data=request.data)


    if serializer.is_valid():
        serializer.save(hospital=hospital)

    else:
        print(serializer.errors) 
    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def getWards(request):
    ward_id = request.data.get("ward")
    hospital_id = request.data.get("hospital")

    if ward_id is not None:
        ward_uuid = uuid.UUID(ward_id)
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        ward = Ward.objects.get(hospital=hospital, ward_id=ward_uuid)

        serializer = WardSerializer(ward)
        return Response(serializer.data)

    else:
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        wards = Ward.objects.filter(hospital=hospital)

        serializer = WardSerializer(wards, many=True)
        return Response(serializer.data)

csrf_exempt
@api_view(['PUT'])
def updateWard(request):
    ward_id = request.data.get("ward")
    hospital = request.data.get("hospital")
    ward_uuid = uuid.UUID(ward_id)
    hospital_uuid = uuid.UUID(hospital)
    hospital_id = get_object_or_404(Company, company_id=hospital_uuid)
    ward = Ward.objects.get(hospital=hospital_id,ward_id=ward_uuid)
    
    serializer = WardSerializer(ward, data=request.data)

    if serializer.is_valid():
        serializer.save()

    else:
        print(serializer.errors)

    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def deleteWard(request):
    ward_id = request.data.get("ward")
    hospital = request.data.get("hospital")
    hospital_uuid = uuid.UUID(hospital)
    ward_uuid = uuid.UUID(ward_id)
    ward = Ward.objects.get(hospital=hospital_uuid, ward_id=ward_uuid)

    ward.delete() 
    message = {'msg':"The ward has been succesfully deleted"} 
    return Response(message)

        #WARDS REPORTS

@csrf_exempt 
@api_view(['POST'])
def generate_wards_pdf(request):
    wardsList = []
    data = json.loads(request.body)
    ward_code = data['ward_code']
    ward_name = data['ward_name']
    wing = data['wing']
    category = data['category']
    hospital_id = data['hospital']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_wards = Ward.objects.filter(hospital=hospital_uuid)

    wards = hospital_wards.filter(Q(ward_code__icontains=ward_code) & Q(ward_name__icontains=ward_name)& Q(wing__icontains=wing))

    if category:
        wards = wards.filter(category = category)

    for wrd in wards:
        obj = {
            "ward_id": wrd.ward_id,
            "ward_code": wrd.ward_code,
            "ward_name": wrd.ward_name,
            "wing": wrd.wing,
            "category": wrd.category,
        }
        wardsList.append(obj)

    context = {"wards":wardsList}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/rooms/templates/rooms')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('wardsPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Wards.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/rooms/static/rooms/roomsPDF.css")

    path = 'Wards.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Wards.pdf'
    pdf.close()
    os.remove("Wards.pdf")  # remove the locally created pdf file.
    return response

@api_view(['POST'])
@csrf_exempt
def generate_wards_excel(request):
    wardsList = []
    data = json.loads(request.body)
    ward_code = data['ward_code']
    ward_name = data['ward_name']
    wing = data['wing']
    category = data['category']
    hospital_id = data['hospital']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_wards = Ward.objects.filter(hospital=hospital_uuid)

    wards = hospital_wards.filter(Q(ward_code__icontains=ward_code) & Q(ward_name__icontains=ward_name)& Q(wing__icontains=wing))

    if category:
        wards = wards.filter(category = category)

    for wrd in wards:
        obj = {
            "ward_id": wrd.ward_id,
            "ward_code": wrd.ward_code,
            "ward_name": wrd.ward_name,
            "wing": wrd.wing,
            "category": wrd.category,
        }
        wardsList.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Wards.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Wards")

    row_num = 0
    columns = ['Code','Name','Wing','Category']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style = style1)

    for wrd in wardsList:
        row_num += 1
        row = [wrd['ward_code'],wrd['ward_name'], wrd['wing'], wrd['category']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@api_view(['POST'])
@csrf_exempt
def generate_wards_csv(request):
    wardsList = []
    data = json.loads(request.body)
    ward_code = data['ward_code']
    ward_name = data['ward_name']
    wing = data['wing']
    category = data['category']
    hospital_id = data['hospital']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_wards = Ward.objects.filter(hospital=hospital_uuid)

    wards = hospital_wards.filter(Q(ward_code__icontains=ward_code) & Q(ward_name__icontains=ward_name)& Q(wing__icontains=wing))

    if category:
        wards = wards.filter(category = category)

    for wrd in wards:
        obj = {
            "ward_id": wrd.ward_id,
            "ward_code": wrd.ward_code,
            "ward_name": wrd.ward_name,
            "wing": wrd.wing,
            "category": wrd.category,
        }
        wardsList.append(obj)



    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Wards.csv'

    writer = csv.writer(response)
    writer.writerow(['Code','Name','Wing','Category'])

    for wrd in wardsList:
        writer.writerow([wrd['ward_code'],wrd['ward_name'], wrd['wing'], wrd['category']])
    return response

@csrf_exempt
@api_view(['POST'])
def display_wards_import_excel(request):
    wardsList = []
    excel_file = request.FILES['wards_excel']
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        code,name, wing,category = row
        obj = {
            "ward_code": code,
            "ward_name": name,
            "wing": wing,
            "category": category
        }
        wardsList.append(obj)

    return JsonResponse({"wards": wardsList})


@csrf_exempt
@api_view(['POST'])
def import_wards_excel(request):

    excel_file = request.FILES['wards_excel']
    hospital_id = request.data.get('hospital')
    hospital_uuid = uuid.UUID(hospital_id)
    hospital = get_object_or_404(Company, company_id=hospital_uuid)

    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        code,name, wing,category = row
        Ward.objects.create(ward_code=code, ward_name=name, wing=wing, hospital=hospital, category=category) 
        

    return HttpResponse("Excel Import Succesful")


        #BEDS VIEWS
    
class BedList(generics.ListCreateAPIView):
    queryset = Bed.objects.all()
    serializer_class = BedSerializer

class BedDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = Bed.objects.all()
    serializer_class = BedSerializer


@csrf_exempt
@api_view(['POST'])
def createBed(request):
    hospital_id = request.data.get("hospital")
    ward_id = request.data.get("ward")
    hospital_uuid = uuid.UUID(hospital_id)
    ward_uuid = uuid.UUID(ward_id)
    hospital = get_object_or_404(Company, company_id=hospital_uuid)
    ward = Ward.objects.get(ward_id=ward_uuid)
    serializer = BedSerializer(data=request.data)


    if serializer.is_valid():
        serializer.save(ward=ward,hospital=hospital)

    else:
        print(serializer.errors) 
    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def getWardBeds(request):
    bed_id = request.data.get("bed")
    hospital_id = request.data.get("hospital")
    ward_id = request.data.get("ward")

    if bed_id is not None and ward_id is not None:
        ward_uuid = uuid.UUID(ward_id)
        hospital_uuid = uuid.UUID(hospital_id)
        bed_uuid = uuid.UUID(bed_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        bed = Bed.objects.get(hospital=hospital, ward=ward_uuid, bed_id=bed_uuid)

        serializer = BedSerializer(bed)
        return Response(serializer.data)

    elif ward_id is not None:
        ward_uuid = uuid.UUID(ward_id)
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        bed = Bed.objects.get(hospital=hospital, ward=ward_uuid)

        serializer = BedSerializer(bed, many=True)
        return Response(serializer.data)

    else:
        hospital_uuid = uuid.UUID(hospital_id)
        hospital = get_object_or_404(Company, company_id=hospital_uuid)
        beds = Bed.objects.filter(hospital=hospital)

        serializer = BedSerializer(beds, many=True)
        return Response(serializer.data)

csrf_exempt
@api_view(['PUT'])
def updateWardBed(request):
    bed_id = request.data.get("bed")
    hospital = request.data.get("hospital")
    bed_uuid = uuid.UUID(bed_id)
    hospital_uuid = uuid.UUID(hospital)
    hospital_id = get_object_or_404(Company, company_id=hospital_uuid)
    bed = Bed.objects.get(hospital=hospital_id,bed_id=bed_uuid)
    
    serializer = BedSerializer(bed, data=request.data)

    if serializer.is_valid():
        serializer.save()

    else:
        print(serializer.errors)

    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def deleteWardBed(request):
    bed_id = request.data.get("bed")
    ward_id = request.data.get("ward")
    hospital = request.data.get("hospital")
    hospital_uuid = uuid.UUID(hospital)
    ward_uuid = uuid.UUID(ward_id)
    bed_uuid = uuid.UUID(bed_id)
    ward = get_object_or_404(Ward, hospital=hospital_uuid, ward_id=ward_uuid)
    bed = Bed.objects.get(hospital=hospital_uuid,ward=ward, bed_id=bed_uuid)

    bed.delete() 
    message = {'msg':"The bed has been succesfully deleted"} 
    return Response(message)

        #BED REPORTS

@csrf_exempt 
@api_view(['POST'])
def generate_beds_pdf(request):
    bedsList = []
    empty = ""
    data = json.loads(request.body)
    bed_number = data['bed_number']
    ward = data['ward']
    status = data['status']
    patient = data['patient']
    price = data['price']
    hospital_id = data['hospital']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_beds = Bed.objects.filter(hospital=hospital_uuid)

    beds = hospital_beds.filter(Q(bed_number__icontains=bed_number) & (Q(patient__first_name__icontains=patient)|Q(patient__last_name__icontains=patient))
                                & Q(ward__ward_name__icontains=ward))

    if status:
        beds = beds.filter(status = status)

    for bed in beds:
        if(patient):
            obj = {
                "bed_id": bed.bed_id,
                "bed_number": bed.bed_number,
                "status": bed.status,
                "ward_id": bed.ward.ward_id,
                "ward_name": bed.ward.ward_name,
                "price": bed.price,
                "patient_id": bed.patient.patient_id,
                "patient_name": bed.patient.first_name + " "+bed.patient.last_name,
            }
            bedsList.append(obj)
        else:
            obj = {
                "bed_id": bed.bed_id,
                "bed_number": bed.bed_number,
                "status": bed.status,
                "ward_id": bed.ward.ward_id,
                "ward_name": bed.ward.ward_name,
                "price": bed.price,
                "patient_id": empty,
                "patient_name": empty,
            }
            bedsList.append(obj)

    context = {"beds":bedsList}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/rooms/templates/rooms')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('bedsPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Beds.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/rooms/static/rooms/roomsPDF.css")

    path = 'Beds.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Beds.pdf'
    pdf.close()
    os.remove("Beds.pdf")  # remove the locally created pdf file.
    return response

@api_view(['POST'])
@csrf_exempt
def generate_beds_excel(request):
    bedsList = []
    empty = ""
    data = json.loads(request.body)
    bed_number = data['bed_number']
    ward = data['ward']
    status = data['status']
    patient = data['patient']
    price = data['price']
    hospital_id = data['hospital']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_beds = Bed.objects.filter(hospital=hospital_uuid)

    beds = hospital_beds.filter(Q(bed_number__icontains=bed_number) & (Q(patient__first_name__icontains=patient)|Q(patient__last_name__icontains=patient))
                                & Q(ward__ward_name__icontains=ward))

    if status:
        beds = beds.filter(status = status)

    for bed in beds:
        if(patient):
            obj = {
                "bed_id": bed.bed_id,
                "bed_number": bed.bed_number,
                "status": bed.status,
                "ward_id": bed.ward.ward_id,
                "ward_name": bed.ward.ward_name,
                "price": bed.price,
                "patient_id": bed.patient.patient_id,
                "patient_name": bed.patient.first_name + " "+bed.patient.last_name,
            }
            bedsList.append(obj)
        else:
            obj = {
                "bed_id": bed.bed_id,
                "bed_number": bed.bed_number,
                "status": bed.status,
                "ward_id": bed.ward.ward_id,
                "ward_name": bed.ward.ward_name,
                "price": bed.price,
                "patient_id": empty,
                "patient_name": empty,
            }
            bedsList.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Beds.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Beds")

    row_num = 0
    columns = ['Bed Number','Ward Name','Price','Status','Patient']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style = style1)

    for bed in bedsList:
        row_num += 1
        row = [bed['bed_number'],bed['ward_name'], bed['price'], bed['status'], bed['patient_name']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@api_view(['POST'])
@csrf_exempt
def generate_beds_csv(request):
    bedsList = []
    empty = ""
    data = json.loads(request.body)
    bed_number = data['bed_number']
    ward = data['ward']
    status = data['status']
    patient = data['patient']
    price = data['price']
    hospital_id = data['hospital']

    hospital_uuid = uuid.UUID(hospital_id)
    hospital_beds = Bed.objects.filter(hospital=hospital_uuid)

    beds = hospital_beds.filter(Q(bed_number__icontains=bed_number) & (Q(patient__first_name__icontains=patient)|Q(patient__last_name__icontains=patient))
                                & Q(ward__ward_name__icontains=ward))

    if status:
        beds = beds.filter(status = status)

    for bed in beds:
        if(patient):
            obj = {
                "bed_id": bed.bed_id,
                "bed_number": bed.bed_number,
                "status": bed.status,
                "ward_id": bed.ward.ward_id,
                "ward_name": bed.ward.ward_name,
                "price": bed.price,
                "patient_id": bed.patient.patient_id,
                "patient_name": bed.patient.first_name + " "+bed.patient.last_name,
            }
            bedsList.append(obj)
        else:
            obj = {
                "bed_id": bed.bed_id,
                "bed_number": bed.bed_number,
                "status": bed.status,
                "ward_id": bed.ward.ward_id,
                "ward_name": bed.ward.ward_name,
                "price": bed.price,
                "patient_id": empty,
                "patient_name": empty,
            }
            bedsList.append(obj)


    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Beds.csv'

    writer = csv.writer(response)
    writer.writerow(['Bed Number','Ward Name','Price','Status','Patient'])

    for bed in bedsList:
        writer.writerow([bed['bed_number'],bed['ward_name'], bed['price'], bed['status'], bed['patient_name']])
    return response

@csrf_exempt
@api_view(['POST'])
def display_beds_import_excel(request):
    bedsList = []
    excel_file = request.FILES['beds_excel']
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        bed_number,ward_name, price,status, patient = row
        obj = {
            "bed_number": bed_number,
            "ward_name": ward_name,
            "price": price,
            "status": status,
            "patient_name": patient,
        }
        bedsList.append(obj)

    return JsonResponse({"beds": bedsList})


@csrf_exempt
@api_view(['POST'])
def import_beds_excel(request):

    excel_file = request.FILES['beds_excel']
    hospital_id = request.data.get('hospital')
    hospital_uuid = uuid.UUID(hospital_id)
    hospital = get_object_or_404(Company, company_id=hospital_uuid)

    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        bed_number,ward_name, price,status, patient = row
        ward_id = Ward.objects.get(name=ward_name,hospital=hospital)
        Bed.objects.create(bed_number=bed_number, ward=ward_id, price=price, status=status, hospital=hospital) 
        

    return HttpResponse("Excel Import Succesful")