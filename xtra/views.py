import os
from django.shortcuts import render,get_object_or_404
from django.http import HttpResponse, JsonResponse
from .models import *
from users.models import Manager
from company.models import Company
from .serializers import *
from rest_framework.views import APIView
from rest_framework.decorators import api_view
from rest_framework import generics,status
from rest_framework.response import Response
#Pagination
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
#Pdf 
import jinja2
import pdfkit
#Excel
import xlwt
#CSV
import csv

import json
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt

import uuid

#Data Import
from openpyxl import load_workbook

# Create your views here.

        #PAGINATION

class BasePagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 1000

class DefaultPagination(PageNumberPagination):
    page_size = 1000
    page_size_query_param = 'page_size'
    max_page_size = 1000

            #DEPARTMENTS VIEWS
    
class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    pagination_class = BasePagination

class DepartmentList(generics.ListCreateAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    pagination_class = DefaultPagination

class DepartmentDetails(generics.RetrieveUpdateDestroyAPIView):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
        

@csrf_exempt
@api_view(['POST'])
def createDepartment(request):
    company_id = request.data.get("company")
    company_uuid = uuid.UUID(company_id)
    company = get_object_or_404(Company, company_id=company_uuid)
    serializer = DepartmentSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save(company=company)

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)


@csrf_exempt
@api_view(['POST'])
def getDepartments(request):
    department_id = request.data.get("department")
    company_id = request.data.get("company")

    if department_id is not None:
        company_uuid = uuid.UUID(company_id)
        department_uuid = uuid.UUID(department_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        department = Department.objects.get(company=company, department_id=department_uuid)

        serializer = DepartmentSerializer(department)
        return Response(serializer.data)

    else:
        company_uuid = uuid.UUID(company_id)
        company = get_object_or_404(Company, company_id=company_uuid)
        department = Department.objects.filter(company=company)

        serializer = DepartmentSerializer(department, many=True)
        return Response(serializer.data)


@csrf_exempt
@api_view(['PUT'])
def updateDepartment(request):
    department_id = request.data.get("department")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    department_uuid = uuid.UUID(department_id)
    department = Department.objects.get(company=company_id,department_id=department_uuid)
    
    serializer = DepartmentSerializer(department, data=request.data)

    if serializer.is_valid():
        serializer.save()

    else:
        print(serializer.errors) 

    
    return Response(serializer.data)

@csrf_exempt
@api_view(['POST'])
def deleteDepartment(request):
    department_id = request.data.get("department")
    company = request.data.get("company")
    company_uuid = uuid.UUID(company)
    company_id = get_object_or_404(Company, company_id=company_uuid)
    department_uuid = uuid.UUID(department_id)
    department = Department.objects.get(company=company_id,department_id=department_uuid)

    department.delete() 
    message = {'msg':"The department has been succesfully deleted"} 
    return Response(message)


@csrf_exempt   
def generate_departments_pdf(request):
    departments = []
    empty = ""
    data = json.loads(request.body)
    code = data['code']
    name = data['name']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_departments = Department.objects.filter(company=company_uuid)

    departList = company_departments.filter(Q(code__icontains=code) & Q(name__icontains=name) )

    for dep in departList:
        manager = Manager.objects.filter(department=dep, status="Active")
        if len(manager):
            obj = {
                "department_id": dep.department_id,
                "code": dep.code,
                "name": dep.name,
                "manager_first_name": manager[0].user.first_name,
                "manager_last_name": manager[0].user.last_name,
                "start_date": manager[0].start_date.strftime("%d %b, %Y")
            }
            departments.append(obj)
        else:
            obj = {
                "department_id": dep.department_id,
                "code": dep.code,
                "name": dep.name,
                "manager_first_name": empty,
                "manager_last_name": empty,
                "start_date": empty
            }
            departments.append(obj)

    context = {"departments":departments}

    template_loader = jinja2.FileSystemLoader('/home/sammyb/Hospital Management System/hms/doctor_profile/templates/doctor_profile')
    template_env = jinja2.Environment(loader=template_loader)

    template  = template_env.get_template('departmentPDF.html')
    output_text = template.render(context)

    config = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    options={"enable-local-file-access": None,
             }

    pdfkit.from_string(output_text, 'Departments.pdf', configuration=config, options=options, css="/home/sammyb/Hospital Management System/hms/doctor_profile/static/doctor_profile/departmentPDF.css")

    path = 'Departments.pdf'
    with open(path, 'rb') as pdf:
        contents = pdf.read()

    response = HttpResponse(contents, content_type='application/pdf')

    response['Content-Disposition'] = 'attachment; filename=Departments.pdf'
    pdf.close()
    os.remove("Departments.pdf")  # remove the locally created pdf file.
    return response

@csrf_exempt
def generate_departments_excel(request):
    departments = []
    empty = ""
    data = json.loads(request.body)
    code = data['code']
    name = data['name']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_departments = Department.objects.filter(company=company_uuid)

    departList = company_departments.filter(Q(code__icontains=code) & Q(name__icontains=name) )

    for dep in departList:
        manager = Manager.objects.filter(department=dep,status="Active")
        if len(manager):
            obj = {
                "department_id": dep.department_id,
                "code": dep.code,
                "name": dep.name,
                "manager_name": manager[0].user.first_name + ' '+ manager[0].user.last_name,
                "start_date": manager[0].start_date.strftime("%d %b, %Y")
            }
            departments.append(obj)
        else:
            obj = {
                "department_id": dep.department_id,
                "code": dep.code,
                "name": dep.name,
                "manager_name": empty,
                "start_date": empty
            }
            departments.append(obj)


    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Departments.xls'

    workbook = xlwt.Workbook()

    worksheet = workbook.add_sheet("Departments")

    row_num = 0
    columns = ['Code', 'Name', 'Manager', 'Start Date']
    style1 = xlwt.easyxf('font:bold 1')
    for col_num in range(len(columns)):
        worksheet.write(row_num, col_num, columns[col_num],style=style1)

    for dep in departments:
        row_num += 1
        row = [dep['code'],dep['name'],dep['manager_name'],dep['start_date']]
        for col_num in range(len(row)):
            worksheet.write(row_num, col_num, row[col_num])
       
    workbook.save(response)
    return response

@csrf_exempt
def generate_departments_csv(request):
    departments = []
    empty = ""
    data = json.loads(request.body)
    code = data['code']
    name = data['name']
    company_id = data['company_id']

    company_uuid = uuid.UUID(company_id)
    company_departments = Department.objects.filter(company=company_uuid)

    departList = company_departments.filter(Q(code__icontains=code) & Q(name__icontains=name) )

    for dep in departList:
        manager = Manager.objects.filter(department=dep, status="Active")
        if len(manager):
            obj = {
                "department_id": dep.department_id,
                "code": dep.code,
                "name": dep.name,
                "manager_name": manager[0].user.first_name + ' '+ manager[0].user.last_name,
                "start_date": manager[0].start_date.strftime("%d %b, %Y")
            }
            departments.append(obj)
        else:
            obj = {
                "department_id": dep.department_id,
                "code": dep.code,
                "name": dep.name,
                "manager_name": empty,
                "start_date": empty
            }
            departments.append(obj)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=Departments.csv'

    writer = csv.writer(response)
    writer.writerow(['Code', 'Name', 'Manager', 'Start Date'])

    for dep in departments:
        writer.writerow([dep['code'],dep['name'],dep['manager_name'],dep['start_date']])
    return response

@csrf_exempt
@api_view(['POST'])
def display_departments_import_excel(request):
    departList = []
    excel_file = request.FILES['departments_excel']
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        code,name = row
        obj = {
            "code": code,
            "name": name
        }
        departList.append(obj)

    return JsonResponse({"departments": departList})


@csrf_exempt
@api_view(['POST'])
def import_departments_excel(request):

    excel_file = request.FILES['departments_excel']
    company_id = request.data.get('company_id')
    company_uuid = uuid.UUID(company_id)
    company = get_object_or_404(Company, company_id=company_uuid)
    departments = Department.objects.filter(company=company)
    wb = load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        code,name = row
        Department.objects.create(code=code, name=name, company=company) 
        

    return HttpResponse("Excel Import Succesful")